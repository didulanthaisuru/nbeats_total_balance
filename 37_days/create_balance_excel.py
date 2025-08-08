import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_balance_excel():
    """
    Create Excel file from shihara_test_37days_version2.xlsx with:
    1. Date vs Balance
    2. Total balance for each particular day
    3. Full date range from start to end date
    4. Fill missing dates with previous day's balance
    """
    
    print("Loading test data file...")
    
    # Load the test data
    try:
        df_test = pd.read_excel("shihara_test_37days_version2.xlsx")
        print(f"✓ Successfully loaded test data with shape: {df_test.shape}")
    except FileNotFoundError:
        print("❌ Error: shihara_test_37days_version2.xlsx not found!")
        return
    except Exception as e:
        print(f"❌ Error loading file: {e}")
        return
    
    # Display basic info about the data
    print(f"\nTest data info:")
    print(f"Columns: {df_test.columns.tolist()}")
    print(f"Data types: {df_test.dtypes}")
    print(f"First few rows:")
    print(df_test.head())
    
    # Extract Date and Balance columns
    print("\nExtracting Date and Balance columns...")
    
    if 'Date' not in df_test.columns or 'Balance' not in df_test.columns:
        print("❌ Error: Required columns 'Date' and 'Balance' not found!")
        print(f"Available columns: {df_test.columns.tolist()}")
        return
    
    # Create a copy with only Date and Balance
    df_balance = df_test[['Date', 'Balance']].copy()
    
    # Convert Date to datetime if not already
    df_balance['Date'] = pd.to_datetime(df_balance['Date'])
    
    # Convert Balance to numeric, handling any non-numeric values
    df_balance['Balance'] = pd.to_numeric(df_balance['Balance'], errors='coerce')
    
    # Remove any rows with NaN values
    initial_count = len(df_balance)
    df_balance = df_balance.dropna()
    final_count = len(df_balance)
    
    if initial_count != final_count:
        print(f"⚠ Removed {initial_count - final_count} rows with missing/invalid data")
    
    print(f"✓ Cleaned data shape: {df_balance.shape}")
    
    # Group by date and take the last balance for each date
    print("\nGrouping by date and taking last balance for each day...")
    df_balance = df_balance.groupby('Date', as_index=False).last()
    df_balance = df_balance.sort_values('Date')
    
    print(f"✓ After grouping: {df_balance.shape} unique dates")
    print(f"Date range: {df_balance['Date'].min()} to {df_balance['Date'].max()}")
    
    # Create full date range from start to end date
    print("\nCreating full date range...")
    start_date = df_balance['Date'].min()
    end_date = df_balance['Date'].max()
    
    # Create complete date range
    full_date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    print(f"✓ Full date range: {len(full_date_range)} days")
    print(f"Original data: {len(df_balance)} days")
    print(f"Missing dates to fill: {len(full_date_range) - len(df_balance)} days")
    
    # Create DataFrame with full date range
    df_full = pd.DataFrame({'Date': full_date_range})
    
    # Merge with original balance data
    df_full = df_full.merge(df_balance, on='Date', how='left')
    
    # Forward fill missing balance values (use previous day's balance)
    df_full['Balance'] = df_full['Balance'].ffill()
    
    # For any remaining NaN values at the beginning, use the first available balance
    if df_full['Balance'].isna().any():
        first_valid_balance = df_full['Balance'].dropna().iloc[0]
        df_full['Balance'] = df_full['Balance'].fillna(first_valid_balance)
    
    print(f"✓ Final dataset shape: {df_full.shape}")
    print(f"Missing values: {df_full['Balance'].isna().sum()}")
    
    # Add additional useful columns
    print("\nAdding additional columns...")
    
    # Day of week
    df_full['Day_of_Week'] = df_full['Date'].dt.day_name()
    
    # Month and Year
    df_full['Month'] = df_full['Date'].dt.month
    df_full['Year'] = df_full['Date'].dt.year
    
    # Balance change from previous day
    df_full['Balance_Change'] = df_full['Balance'].diff()
    
    # Balance change percentage
    df_full['Balance_Change_Percent'] = (df_full['Balance_Change'] / df_full['Balance'].shift(1)) * 100
    
    # Fill NaN values in change columns
    df_full['Balance_Change'] = df_full['Balance_Change'].fillna(0)
    df_full['Balance_Change_Percent'] = df_full['Balance_Change_Percent'].fillna(0)
    
    # Format the data for better display
    df_display = df_full.copy()
    df_display['Date'] = df_display['Date'].dt.strftime('%Y-%m-%d')
    df_display['Balance'] = df_display['Balance'].round(2)
    df_display['Balance_Change'] = df_display['Balance_Change'].round(2)
    df_display['Balance_Change_Percent'] = df_display['Balance_Change_Percent'].round(2)
    
    # Reorder columns for better readability
    column_order = ['Date', 'Day_of_Week', 'Balance', 'Balance_Change', 'Balance_Change_Percent', 'Month', 'Year']
    df_display = df_display[column_order]
    
    print(f"✓ Final display dataset shape: {df_display.shape}")
    
    # Create Excel file with formatting
    print("\nCreating Excel file with formatting...")
    output_filename = "balance_daily_report.xlsx"
    
    # Save to Excel
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_display.to_excel(writer, sheet_name='Daily_Balance', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Daily_Balance']
        
        # Apply formatting
        format_excel_worksheet(worksheet, df_display)
    
    print(f"✓ Excel file created: {output_filename}")
    
    # Display summary statistics
    print("\n📊 SUMMARY STATISTICS:")
    print(f"   • Total days: {len(df_full):,}")
    print(f"   • Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    print(f"   • Balance range: {df_full['Balance'].min():,.2f} to {df_full['Balance'].max():,.2f}")
    print(f"   • Average balance: {df_full['Balance'].mean():,.2f}")
    print(f"   • Days with balance changes: {(df_full['Balance_Change'] != 0).sum():,}")
    print(f"   • Days with no change: {(df_full['Balance_Change'] == 0).sum():,}")
    print(f"   • Largest single-day increase: {df_full['Balance_Change'].max():,.2f}")
    print(f"   • Largest single-day decrease: {df_full['Balance_Change'].min():,.2f}")
    
    # Show sample of the data
    print(f"\n📋 SAMPLE DATA (First 10 rows):")
    print(df_display.head(10).to_string(index=False))
    
    print(f"\n📋 SAMPLE DATA (Last 10 rows):")
    print(df_display.tail(10).to_string(index=False))
    
    return df_display

def format_excel_worksheet(worksheet, df):
    """
    Apply formatting to the Excel worksheet
    """
    # Import required styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format specific columns
    # Date column - left align
    for cell in worksheet['A'][1:]:
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Day of week column - left align
    for cell in worksheet['B'][1:]:
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Balance column - right align with number format
    for cell in worksheet['C'][1:]:
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '#,##0.00'
    
    # Balance change columns - right align with number format
    for cell in worksheet['D'][1:]:
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '#,##0.00'
    
    for cell in worksheet['E'][1:]:
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '0.00%'
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add conditional formatting for balance changes
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    # Green for positive changes
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    worksheet.conditional_formatting.add(
        f'D2:D{worksheet.max_row}',
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
    )
    
    # Red for negative changes
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    worksheet.conditional_formatting.add(
        f'D2:D{worksheet.max_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
    )

if __name__ == "__main__":
    print("=" * 60)
    print("BALANCE DAILY REPORT GENERATOR")
    print("=" * 60)
    
    result_df = create_balance_excel()
    
    if result_df is not None:
        print("\n" + "=" * 60)
        print("✅ SUCCESS: Balance daily report created successfully!")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("❌ FAILED: Could not create balance daily report!")
        print("=" * 60) 